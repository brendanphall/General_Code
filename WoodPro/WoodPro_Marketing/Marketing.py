import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference, Series, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart.series import SeriesLabel


def create_geospatial_pricing_calculator():
    # Create a workbook
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Define color constants
    COLOR_HEADER = "D6E8F6"  # Light blue
    COLOR_INPUT = "FFF2CC"  # Light yellow
    COLOR_CALC = "FFFFFF"  # White
    COLOR_TOTAL = "E6F0F9"  # Light blue for totals
    COLOR_BELOW_TARGET = "FFC7CE"  # Light red
    COLOR_ABOVE_TARGET = "C6EFCE"  # Light green
    COLOR_NAV = "F2F2F2"  # Light gray for navigation

    # Define common styles
    header_font = Font(bold=True)
    money_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Function to create navigation bar
    def add_navigation_bar(sheet):
        sheet_names = ["Client Info", "License & Infrastructure", "Implementation",
                       "Support & Maintenance", "Contract Summary",
                       "Profitability Analysis", "Parameters", "Proposal Generator"]

        for idx, name in enumerate(sheet_names, 1):
            cell = sheet.cell(row=1, column=idx)
            cell.value = name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=COLOR_NAV, end_color=COLOR_NAV, fill_type="solid")
            cell.hyperlink = f"#{name}!A1"
            cell.style = "Hyperlink"

        sheet.row_dimensions[1].height = 20
        sheet.merge_cells('A2:H2')
        separator = sheet.cell(row=2, column=1)
        separator.value = ""
        separator.fill = PatternFill(start_color=COLOR_NAV, end_color=COLOR_NAV, fill_type="solid")

    # Function to create sheet header
    def add_sheet_header(sheet, title):
        sheet.merge_cells('A3:H3')
        header = sheet.cell(row=3, column=1)
        header.value = title
        header.font = Font(bold=True, size=14)
        header.alignment = Alignment(horizontal='center')
        sheet.row_dimensions[3].height = 24

    # ----------------------- Create Parameters Sheet -----------------------

    def create_parameters_sheet():
        params_sheet = wb.create_sheet("Parameters")
        add_navigation_bar(params_sheet)
        add_sheet_header(params_sheet, "Parameters and Standard Rates")

        # Standard Hourly Rates
        params_sheet['A5'] = "Standard Hourly Rates"
        params_sheet['A5'].font = header_font

        hourly_rates_df = pd.DataFrame({
            'Role': ['Project Manager', 'Business Analyst', 'System Architect', 'Developer',
                     'QA Tester', 'Trainer', 'Support Specialist'],
            'Internal Cost': [80, 75, 90, 85, 70, 65, 60],
            'Client Rate': [135, 150, 175, 165, 125, 125, 125]
        })

        for r_idx, row in enumerate(dataframe_to_rows(hourly_rates_df, index=False, header=True), 6):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 6:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                elif c_idx in [2, 3]:  # Money columns
                    cell.number_format = money_format

        # Format the rates table
        for row in range(6, 14):
            for col in range(1, 4):
                params_sheet.cell(row=row, column=col).border = thin_border

        # License Costs
        params_sheet['A15'] = "License Costs"
        params_sheet['A15'].font = header_font

        license_costs_df = pd.DataFrame({
            'License Type': ['ESRI Base Hosted', 'ESRI add-on per 4 Cores', 'Lite Partner - Viewer',
                             'Basic Partner', 'Standard Partner', 'Mobile Partner',
                             'WoodPro - Editor', 'WoodPro - Viewer'],
            'Our Cost': [2200, 2200, 125, 250, 700, 400, 350, 60],
            'Standard Markup %': [0.20, 0.20, 0.25, 0.25, 0.20, 0.25, 0.25, 0.30]
        })

        for r_idx, row in enumerate(dataframe_to_rows(license_costs_df, index=False, header=True), 16):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 16:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                elif c_idx == 2:  # Money column
                    cell.number_format = money_format
                elif c_idx == 3:  # Percentage column
                    cell.number_format = percent_format

        # Format the license costs table
        for row in range(16, 25):
            for col in range(1, 4):
                params_sheet.cell(row=row, column=col).border = thin_border

        # Target Margins
        params_sheet['A26'] = "Target Margins"
        params_sheet['A26'].font = header_font

        target_margins_df = pd.DataFrame({
            'Service Category': ['Software Licenses', 'Infrastructure', 'Implementation',
                                 'Customization', 'Support & Maintenance', 'Overall Solution'],
            'Target Margin %': [0.25, 0.25, 0.45, 0.50, 0.55, 0.40]
        })

        for r_idx, row in enumerate(dataframe_to_rows(target_margins_df, index=False, header=True), 27):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 27:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                elif c_idx == 2:  # Percentage column
                    cell.number_format = percent_format

        # Format the target margins table
        for row in range(27, 34):
            for col in range(1, 3):
                params_sheet.cell(row=row, column=col).border = thin_border

        # Complexity Multipliers
        params_sheet['A35'] = "Complexity Multipliers"
        params_sheet['A35'].font = header_font

        complexity_df = pd.DataFrame({
            'Complexity Level': ['Basic', 'Standard', 'Advanced', 'Enterprise'],
            'Hours Multiplier': [1.0, 1.5, 2.5, 4.0],
            'Support Multiplier': [1.0, 1.25, 1.5, 2.0]
        })

        for r_idx, row in enumerate(dataframe_to_rows(complexity_df, index=False, header=True), 36):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 36:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")

        # Format the complexity table
        for row in range(36, 41):
            for col in range(1, 4):
                params_sheet.cell(row=row, column=col).border = thin_border

        # Contract Term Discounts
        params_sheet['A42'] = "Contract Term Discounts"
        params_sheet['A42'].font = header_font

        contract_terms_df = pd.DataFrame({
            'Contract Length': ['Monthly', '1-Year', '2-Year', '3-Year'],
            'Payment Structure': ['Monthly', 'Annual', 'Annual', 'Annual'],
            'Setup Amortization': ['None', '12 months', '24 months', '36 months'],
            'License Discount': [0.00, 0.05, 0.10, 0.15]
        })

        for r_idx, row in enumerate(dataframe_to_rows(contract_terms_df, index=False, header=True), 43):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 43:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                elif c_idx == 4:  # Percentage column
                    cell.number_format = percent_format

        # Format the contract terms table
        for row in range(43, 48):
            for col in range(1, 5):
                params_sheet.cell(row=row, column=col).border = thin_border

        # Support Tiers
        params_sheet['A49'] = "Support Tiers"
        params_sheet['A49'].font = header_font

        support_tiers_df = pd.DataFrame({
            'Support Level': ['Basic', 'Standard', 'Premium', 'Custom SLA'],
            'Description': ['Email support', 'Email + phone', '24/7 support', 'Custom agreement'],
            '% of License Cost': [0.12, 0.18, 0.25, 'Negotiated'],
            'Response Time': ['24 hours', '8 hours', '4 hours', 'Varies']
        })

        for r_idx, row in enumerate(dataframe_to_rows(support_tiers_df, index=False, header=True), 50):
            for c_idx, value in enumerate(row, 1):
                cell = params_sheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 50:  # Header row
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                elif c_idx == 3 and isinstance(value, (int, float)):  # Percentage column
                    cell.number_format = percent_format

        # Format the support tiers table
        for row in range(50, 55):
            for col in range(1, 5):
                params_sheet.cell(row=row, column=col).border = thin_border

        # Adjust column widths in Parameters sheet
        for col in params_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            params_sheet.column_dimensions[column].width = adjusted_width



        # Create named ranges
        wb.defined_names.add(DefinedName(name="HourlyRates", attr_text="Parameters!$B$7:$C$13"))
        wb.defined_names.add(DefinedName(name="LicenseCosts", attr_text="Parameters!$B$17:$C$24"))
        wb.defined_names.add(DefinedName(name="LicenseMarkups", attr_text="Parameters!$C$17:$C$24"))
        wb.defined_names.add(DefinedName(name="TargetMargins", attr_text="Parameters!$B$28:$B$33"))
        wb.defined_names.add(DefinedName(name="ComplexityMultipliers", attr_text="Parameters!$B$37:$C$40"))
        wb.defined_names.add(DefinedName(name="ContractTerms", attr_text="Parameters!$A$44:$D$47"))
        wb.defined_names.add(DefinedName(name="SupportTiers", attr_text="Parameters!$A$51:$D$54"))

        return params_sheet

    # ----------------------- Create Client Info Sheet -----------------------

    def create_client_info_sheet():
        client_sheet = wb.create_sheet("Client Info")
        add_navigation_bar(client_sheet)
        add_sheet_header(client_sheet, "Client Information and Project Overview")

        # Client Information Section
        client_sheet['A5'] = "Client Information"
        client_sheet['A5'].font = header_font

        # Create basic client info fields
        client_info_fields = [
            ('Client Name', ''),
            ('Industry', ''),
            ('Contact Person', ''),
            ('Contact Email', ''),
            ('Contact Phone', '')
        ]

        for idx, (field, default) in enumerate(client_info_fields, 6):
            client_sheet[f'A{idx}'] = field
            client_sheet[f'B{idx}'] = default
            client_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                       fill_type="solid")

        # Project Parameters Section
        client_sheet['A12'] = "Project Parameters"
        client_sheet['A12'].font = header_font

        # Create project parameter fields with dropdowns
        project_params = [
            ('Contract Term', '1-Year', 'ContractTerm'),
            ('Client Size', 'Medium', 'ClientSize'),
            ('Implementation Complexity', 'Standard', 'Complexity'),
            ('Support Tier', 'Standard', 'SupportTier')
        ]

        # Add data validation for dropdowns
        contract_dv = DataValidation(type="list", formula1='"Monthly,1-Year,2-Year,3-Year"', allow_blank=True)
        size_dv = DataValidation(type="list", formula1='"Small,Medium,Large,Enterprise"', allow_blank=True)
        complexity_dv = DataValidation(type="list", formula1='"Basic,Standard,Advanced,Enterprise"', allow_blank=True)
        support_dv = DataValidation(type="list", formula1='"Basic,Standard,Premium,Custom SLA"', allow_blank=True)

        client_sheet.add_data_validation(contract_dv)
        client_sheet.add_data_validation(size_dv)
        client_sheet.add_data_validation(complexity_dv)
        client_sheet.add_data_validation(support_dv)

        for idx, (field, default, validation) in enumerate(project_params, 13):
            client_sheet[f'A{idx}'] = field
            client_sheet[f'B{idx}'] = default
            client_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                       fill_type="solid")

            # Add validation to the appropriate cell
            if validation == 'ContractTerm':
                contract_dv.add(client_sheet[f'B{idx}'])
            elif validation == 'ClientSize':
                size_dv.add(client_sheet[f'B{idx}'])
            elif validation == 'Complexity':
                complexity_dv.add(client_sheet[f'B{idx}'])
            elif validation == 'SupportTier':
                support_dv.add(client_sheet[f'B{idx}'])

        # User Counts Section
        client_sheet['A18'] = "User Counts"
        client_sheet['A18'].font = header_font

        user_types = [
            ('Lite Partner - Viewer', 0),
            ('Basic Partner - Contributor', 0),
            ('Standard Partner - Creator', 0),
            ('Mobile Partner - Mobile Worker', 0),
            ('WoodPro - Editor', 0),
            ('WoodPro - Viewer', 0),
            ('Total Users', '=SUM(B19:B24)')
        ]

        for idx, (user_type, default) in enumerate(user_types, 19):
            client_sheet[f'A{idx}'] = user_type

            if isinstance(default, str) and default.startswith('='):
                client_sheet[f'B{idx}'] = default
                client_sheet[f'B{idx}'].font = Font(bold=True)
            else:
                client_sheet[f'B{idx}'] = default
                client_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                           fill_type="solid")

        # Infrastructure Requirements
        client_sheet['A27'] = "Infrastructure Requirements"
        client_sheet['A27'].font = header_font

        infra_fields = [
            ('ESRI Base', 1),
            ('Additional ESRI Cores', 0),
            ('Storage (GB)', 100),
            ('Backup Frequency', 'Daily')
        ]

        for idx, (field, default) in enumerate(infra_fields, 28):
            client_sheet[f'A{idx}'] = field
            client_sheet[f'B{idx}'] = default
            client_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                       fill_type="solid")

        # Summary Section
        client_sheet['D5'] = "Project Summary"
        client_sheet['D5'].font = header_font

        summary_fields = [
            ('Estimated Implementation Timeline',
             '=IF(B15="Basic","4-6 weeks",IF(B15="Standard","8-12 weeks",IF(B15="Advanced","12-16 weeks","16-24 weeks")))'),
            ('Estimated One-Time Cost', '=\'Contract Summary\'!F16'),
            ('Estimated Monthly Recurring', '=\'Contract Summary\'!F23'),
            ('Estimated Total Contract Value', '=\'Contract Summary\'!F29'),
            ('Estimated Breakeven Point', '=\'Contract Summary\'!C40&" months"'),
            ('Overall Profit Margin', '=\'Contract Summary\'!F30'),
        ]

        for idx, (field, formula) in enumerate(summary_fields, 6):
            client_sheet[f'D{idx}'] = field
            client_sheet[f'E{idx}'] = formula
            if idx == 6:  # Implementation timeline (text)
                pass
            elif idx in [7, 8, 9]:  # Money fields
                client_sheet[f'E{idx}'].number_format = money_format
            elif idx == 11:  # Percentage
                client_sheet[f'E{idx}'].number_format = percent_format

        # Key Metrics Section
        client_sheet['D12'] = "Key Metrics"
        client_sheet['D12'].font = header_font

        metrics_fields = [
            ('Average Cost Per User', '=IF(B25=0,0,\'Contract Summary\'!F23/B25)'),
            ('License Costs as % of Total', '=\'Profitability Analysis\'!E7/\'Profitability Analysis\'!E11'),
            ('Implementation as % of Year 1',
             '=\'Contract Summary\'!F16/(\'Contract Summary\'!F16+\'Contract Summary\'!F23*12)'),
            ('Support as % of Recurring', '=\'Support & Maintenance\'!F15/\'Contract Summary\'!F23'),
        ]

        for idx, (field, formula) in enumerate(metrics_fields, 13):
            client_sheet[f'D{idx}'] = field
            client_sheet[f'E{idx}'] = formula
            if idx == 13:  # Cost per user (money)
                client_sheet[f'E{idx}'].number_format = money_format
            else:  # Percentages
                client_sheet[f'E{idx}'].number_format = percent_format

        # Placeholder for chart
        client_sheet['D18'] = "Revenue Breakdown"
        client_sheet['D18'].font = header_font

        # Adjust column widths
        for col in client_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            client_sheet.column_dimensions[column].width = adjusted_width

        return client_sheet

    # ----------------------- Create License & Infrastructure Sheet -----------------------

    def create_license_sheet():
        license_sheet = wb.create_sheet("License & Infrastructure")
        add_navigation_bar(license_sheet)
        add_sheet_header(license_sheet, "License and Infrastructure Costs")

        # Base Platform Section
        license_sheet['A5'] = "Base Platform"
        license_sheet['A5'].font = header_font

        # Set up column headers
        headers = ['Component', 'Quantity', 'Unit Cost', 'Our Cost', 'Markup %', 'Client Price', 'Monthly Revenue']
        for idx, header in enumerate(headers, 1):
            cell = license_sheet.cell(row=6, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Base platform components
        base_components = [
            ('ESRI Base Hosted Solution', '=\'Client Info\'!B28',
             '=VLOOKUP("ESRI Base Hosted",Parameters!A17:C24,2,FALSE)'),
            ('ESRI add-on per 4 Cores', '=\'Client Info\'!B29',
             '=VLOOKUP("ESRI add-on per 4 Cores",Parameters!A17:C24,2,FALSE)'),
            ('Subtotal', '', '')
        ]

        for idx, (component, qty_formula, cost_formula) in enumerate(base_components, 7):
            license_sheet[f'A{idx}'] = component

            if idx < 9:  # Regular rows
                license_sheet[f'B{idx}'] = qty_formula
                license_sheet[f'C{idx}'] = cost_formula
                license_sheet[f'D{idx}'] = f'=B{idx}*C{idx}'
                license_sheet[f'E{idx}'] = f'=VLOOKUP("{component}",Parameters!A17:C24,3,FALSE)'
                license_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
                license_sheet[f'G{idx}'] = f'=F{idx}/12'
            else:  # Subtotal row
                license_sheet[f'D{idx}'] = '=SUM(D7:D8)'
                license_sheet[f'F{idx}'] = '=SUM(F7:F8)'
                license_sheet[f'G{idx}'] = '=SUM(G7:G8)'
                license_sheet[f'A{idx}'].font = Font(bold=True)
                license_sheet[f'D{idx}'].font = Font(bold=True)
                license_sheet[f'F{idx}'].font = Font(bold=True)
                license_sheet[f'G{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 9 or idx == 9:
                license_sheet[f'C{idx}'].number_format = money_format
                license_sheet[f'D{idx}'].number_format = money_format
                license_sheet[f'F{idx}'].number_format = money_format
                license_sheet[f'G{idx}'].number_format = money_format

            if idx < 9:
                license_sheet[f'E{idx}'].number_format = percent_format

        # User Licenses Section
        license_sheet['A11'] = "User Licenses"
        license_sheet['A11'].font = header_font

        # User license rows
        user_components = [
            ('Lite Partner - Viewer', '=\'Client Info\'!B19',
             '=VLOOKUP("Lite Partner - Viewer",Parameters!A17:C24,2,FALSE)'),
            ('Basic Partner - Contributor', '=\'Client Info\'!B20',
             '=VLOOKUP("Basic Partner",Parameters!A17:C24,2,FALSE)'),
            ('Standard Partner - Creator', '=\'Client Info\'!B21',
             '=VLOOKUP("Standard Partner",Parameters!A17:C24,2,FALSE)'),
            ('Mobile Partner - Mobile Worker', '=\'Client Info\'!B22',
             '=VLOOKUP("Mobile Partner",Parameters!A17:C24,2,FALSE)'),
            ('WoodPro - Editor', '=\'Client Info\'!B23', '=VLOOKUP("WoodPro - Editor",Parameters!A17:C24,2,FALSE)'),
            ('WoodPro - Viewer', '=\'Client Info\'!B24', '=VLOOKUP("WoodPro - Viewer",Parameters!A17:C24,2,FALSE)'),
            ('Subtotal', '', '')
        ]

        # Add column headers again
        for idx, header in enumerate(headers, 1):
            cell = license_sheet.cell(row=12, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        for idx, (component, qty_formula, cost_formula) in enumerate(user_components, 13):
            license_sheet[f'A{idx}'] = component

            if idx < 19:  # Regular rows
                license_sheet[f'B{idx}'] = qty_formula
                license_sheet[f'C{idx}'] = cost_formula
                license_sheet[f'D{idx}'] = f'=B{idx}*C{idx}'
                if component == 'WoodPro - Editor' or component == 'WoodPro - Viewer':
                    license_sheet[f'E{idx}'] = f'=VLOOKUP("{component}",Parameters!A17:C24,3,FALSE)'
                else:
                    name_to_lookup = component.split(' - ')[0]
                    license_sheet[f'E{idx}'] = f'=VLOOKUP("{name_to_lookup}",Parameters!A17:C24,3,FALSE)'
                license_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
                license_sheet[f'G{idx}'] = f'=F{idx}/12'
            else:  # Subtotal row
                license_sheet[f'D{idx}'] = '=SUM(D13:D18)'
                license_sheet[f'F{idx}'] = '=SUM(F13:F18)'
                license_sheet[f'G{idx}'] = '=SUM(G13:G18)'
                license_sheet[f'A{idx}'].font = Font(bold=True)
                license_sheet[f'D{idx}'].font = Font(bold=True)
                license_sheet[f'F{idx}'].font = Font(bold=True)
                license_sheet[f'G{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 19 or idx == 19:
                license_sheet[f'C{idx}'].number_format = money_format
                license_sheet[f'D{idx}'].number_format = money_format
                license_sheet[f'F{idx}'].number_format = money_format
                license_sheet[f'G{idx}'].number_format = money_format

            if idx < 19:
                license_sheet[f'E{idx}'].number_format = percent_format

        # Infrastructure Section
        license_sheet['A21'] = "Infrastructure"
        license_sheet['A21'].font = header_font

        # Set up column headers - slightly different for infrastructure
        infra_headers = ['Component', 'Specification', 'Our Cost', 'Markup %', 'Client Price', 'Monthly Revenue']
        for idx, header in enumerate(infra_headers, 1):
            col_idx = idx if idx < 3 else idx + 1  # Skip the Unit Cost column
            cell = license_sheet.cell(row=22, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Infrastructure components
        infra_components = [
            ('AWS Hosting', '=\'Client Info\'!B25&" users"', '=\'Client Info\'!B25*5*12'),
            ('Data Storage', '=\'Client Info\'!B30&" GB"', '=\'Client Info\'!B30*0.5*12'),
            ('Backup Services', '=\'Client Info\'!B31',
             '=IF(\'Client Info\'!B31="Daily",\'Client Info\'!B30*0.8*12,\'Client Info\'!B30*0.4*12)'),
            ('Security Services', 'Enterprise-grade', '=\'Client Info\'!B25*3*12'),
            ('Subtotal', '', '')
        ]

        for idx, (component, spec_formula, cost_formula) in enumerate(infra_components, 23):
            license_sheet[f'A{idx}'] = component
            license_sheet[f'B{idx}'] = spec_formula

            if idx < 27:  # Regular rows
                license_sheet[f'D{idx}'] = cost_formula
                license_sheet[f'E{idx}'] = 0.25  # 25% markup for infrastructure
                license_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
                license_sheet[f'G{idx}'] = f'=F{idx}/12'
            else:  # Subtotal row
                license_sheet[f'D{idx}'] = '=SUM(D23:D26)'
                license_sheet[f'F{idx}'] = '=SUM(F23:F26)'
                license_sheet[f'G{idx}'] = '=SUM(G23:G26)'
                license_sheet[f'A{idx}'].font = Font(bold=True)
                license_sheet[f'D{idx}'].font = Font(bold=True)
                license_sheet[f'F{idx}'].font = Font(bold=True)
                license_sheet[f'G{idx}'].font = Font(bold=True)

                # Apply number formats
            if idx < 27 or idx == 27:
                license_sheet[f'D{idx}'].number_format = money_format
                license_sheet[f'F{idx}'].number_format = money_format
                license_sheet[f'G{idx}'].number_format = money_format

            if idx < 27:
                license_sheet[f'E{idx}'].number_format = percent_format

                # Total License & Infrastructure
            license_sheet['A29'] = "Total License & Infrastructure"
            license_sheet['A29'].font = header_font

            license_sheet['A30'] = "Total Annual Cost"
            license_sheet['D30'] = '=D9+D19+D27'
            license_sheet['D30'].number_format = money_format
            license_sheet['D30'].font = Font(bold=True)

            license_sheet['F30'] = '=F9+F19+F27'
            license_sheet['F30'].number_format = money_format
            license_sheet['F30'].font = Font(bold=True)

            license_sheet['A31'] = "Total Monthly Cost"
            license_sheet['G31'] = '=G9+G19+G27'
            license_sheet['G31'].number_format = money_format
            license_sheet['G31'].font = Font(bold=True)

            # Apply conditional formatting to margins
            license_sheet_margin_rule = FormulaRule(
                formula=['$F$30/$D$30-1<Parameters!$B$28'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
            )
            license_sheet.conditional_formatting.add('F30:F30', license_sheet_margin_rule)

            # Adjust column widths
            for col in license_sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                license_sheet.column_dimensions[column].width = adjusted_width

            return license_sheet

            # ----------------------- Create Implementation Sheet -----------------------

    def create_implementation_sheet():
        impl_sheet = wb.create_sheet("Implementation")
        add_navigation_bar(impl_sheet)
        add_sheet_header(impl_sheet, "Implementation and Customization")

        # Project Setup Section
        impl_sheet['A5'] = "Project Setup"
        impl_sheet['A5'].font = header_font

        # Set up column headers
        setup_headers = ['Service', 'Hours', 'Hourly Rate', 'Our Cost', 'Markup %', 'Client Price']
        for idx, header in enumerate(setup_headers, 1):
            cell = impl_sheet.cell(row=6, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Project setup tasks
        setup_tasks = [
            ('Project Management', 16, '=VLOOKUP("Project Manager", Parameters!A7:C13, 3, FALSE)'),
            ('Requirements Analysis', 24, '=VLOOKUP("Business Analyst", Parameters!A7:C13, 3, FALSE)'),
            ('Subtotal', '', '')
        ]

        for idx, (task, hours, rate) in enumerate(setup_tasks, 7):
            impl_sheet[f'A{idx}'] = task

            if idx < 9:  # Regular rows
                impl_sheet[f'B{idx}'] = hours
                impl_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                         fill_type="solid")

                impl_sheet[f'C{idx}'] = rate
                impl_sheet[f'D{idx}'] = f'=B{idx}*C{idx}'
                impl_sheet[f'E{idx}'] = 0.50  # 50% markup for services
                impl_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
            else:  # Subtotal row
                impl_sheet[f'D{idx}'] = '=SUM(D7:D8)'
                impl_sheet[f'F{idx}'] = '=SUM(F7:F8)'
                impl_sheet[f'A{idx}'].font = Font(bold=True)
                impl_sheet[f'D{idx}'].font = Font(bold=True)
                impl_sheet[f'F{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 9 or idx == 9:
                impl_sheet[f'C{idx}'].number_format = money_format
                impl_sheet[f'D{idx}'].number_format = money_format
                impl_sheet[f'F{idx}'].number_format = money_format

            if idx < 9:
                impl_sheet[f'E{idx}'].number_format = percent_format

        # System Configuration Section
        impl_sheet['A11'] = "System Configuration"
        impl_sheet['A11'].font = header_font

        # System config headers (same as setup)
        for idx, header in enumerate(setup_headers, 1):
            cell = impl_sheet.cell(row=12, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # System configuration tasks
        config_tasks = [
            ('Basic WoodPro Setup', 20, '=VLOOKUP("System Architect", Parameters!A7:C13, 3, FALSE)'),
            ('User Configuration', 8, '=VLOOKUP("System Architect", Parameters!A7:C13, 3, FALSE)'),
            ('System Integration', 30, '=VLOOKUP("Developer", Parameters!A7:C13, 3, FALSE)'),
            ('Subtotal', '', '')
        ]

        for idx, (task, hours, rate) in enumerate(config_tasks, 13):
            impl_sheet[f'A{idx}'] = task

            if idx < 16:  # Regular rows
                impl_sheet[f'B{idx}'] = f'={hours}*VLOOKUP(\'Client Info\'!B15, Parameters!A37:C40, 2, FALSE)'

                impl_sheet[f'C{idx}'] = rate
                impl_sheet[f'D{idx}'] = f'=B{idx}*C{idx}'
                impl_sheet[f'E{idx}'] = 0.50  # 50% markup for services
                impl_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
            else:  # Subtotal row
                impl_sheet[f'D{idx}'] = '=SUM(D13:D15)'
                impl_sheet[f'F{idx}'] = '=SUM(F13:F15)'
                impl_sheet[f'A{idx}'].font = Font(bold=True)
                impl_sheet[f'D{idx}'].font = Font(bold=True)
                impl_sheet[f'F{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 16 or idx == 16:
                impl_sheet[f'B{idx}'].number_format = '0.0'
                impl_sheet[f'C{idx}'].number_format = money_format
                impl_sheet[f'D{idx}'].number_format = money_format
                impl_sheet[f'F{idx}'].number_format = money_format

            if idx < 16:
                impl_sheet[f'E{idx}'].number_format = percent_format

        # Custom Development Section
        impl_sheet['A18'] = "Custom Development"
        impl_sheet['A18'].font = header_font

        # Custom development headers
        dev_headers = ['Feature', 'Complexity', 'Hours', 'Hourly Rate', 'Our Cost', 'Markup %', 'Client Price']
        for idx, header in enumerate(dev_headers, 1):
            cell = impl_sheet.cell(row=19, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Custom development tasks
        dev_tasks = [
            ('Accounting System Integration', 'Medium', 40, '=VLOOKUP("Developer", Parameters!A7:C13, 3, FALSE)'),
            ('Custom Reports', 'Low', 20, '=VLOOKUP("Developer", Parameters!A7:C13, 3, FALSE)'),
            ('Field Data Collection Workflow', 'High', 50, '=VLOOKUP("Developer", Parameters!A7:C13, 3, FALSE)'),
            ('Subtotal', '', '', '')
        ]

        complexity_dv = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
        impl_sheet.add_data_validation(complexity_dv)

        for idx, (task, complexity, hours, rate) in enumerate(dev_tasks, 20):
            impl_sheet[f'A{idx}'] = task

            if idx < 23:  # Regular rows
                impl_sheet[f'B{idx}'] = complexity
                impl_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                         fill_type="solid")
                complexity_dv.add(impl_sheet[f'B{idx}'])

                impl_sheet[f'C{idx}'] = f'=IF(B{idx}="Low",{hours},IF(B{idx}="Medium",{hours}*1.5,{hours}*2))'
                impl_sheet[f'D{idx}'] = rate
                impl_sheet[f'E{idx}'] = f'=C{idx}*D{idx}'
                impl_sheet[f'F{idx}'] = 0.60  # 60% markup for custom development
                impl_sheet[f'G{idx}'] = f'=E{idx}*(1+F{idx})'
            else:  # Subtotal row
                impl_sheet[f'E{idx}'] = '=SUM(E20:E22)'
                impl_sheet[f'G{idx}'] = '=SUM(G20:G22)'
                impl_sheet[f'A{idx}'].font = Font(bold=True)
                impl_sheet[f'E{idx}'].font = Font(bold=True)
                impl_sheet[f'G{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 23 or idx == 23:
                impl_sheet[f'C{idx}'].number_format = '0.0'
                impl_sheet[f'D{idx}'].number_format = money_format
                impl_sheet[f'E{idx}'].number_format = money_format
                impl_sheet[f'G{idx}'].number_format = money_format

            if idx < 23:
                impl_sheet[f'F{idx}'].number_format = percent_format

        # Data Migration and Training Section
        impl_sheet['A25'] = "Data Migration and Training"
        impl_sheet['A25'].font = header_font

        # Migration headers (same as setup)
        for idx, header in enumerate(setup_headers, 1):
            cell = impl_sheet.cell(row=26, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Data migration and training tasks
        migration_tasks = [
            ('Data Migration', 20, '=VLOOKUP("System Architect", Parameters!A7:C13, 3, FALSE)'),
            ('User Training', 16, '=VLOOKUP("Trainer", Parameters!A7:C13, 3, FALSE)'),
            ('Documentation', 12, '=VLOOKUP("Trainer", Parameters!A7:C13, 3, FALSE)'),
            ('Subtotal', '', '')
        ]

        for idx, (task, hours, rate) in enumerate(migration_tasks, 27):
            impl_sheet[f'A{idx}'] = task

            if idx < 30:  # Regular rows
                impl_sheet[f'B{idx}'] = f'={hours}*VLOOKUP(\'Client Info\'!B15, Parameters!A37:C40, 2, FALSE)'

                impl_sheet[f'C{idx}'] = rate
                impl_sheet[f'D{idx}'] = f'=B{idx}*C{idx}'
                impl_sheet[f'E{idx}'] = 0.45  # 45% markup for training and migration
                impl_sheet[f'F{idx}'] = f'=D{idx}*(1+E{idx})'
            else:  # Subtotal row
                impl_sheet[f'D{idx}'] = '=SUM(D27:D29)'
                impl_sheet[f'F{idx}'] = '=SUM(F27:F29)'
                impl_sheet[f'A{idx}'].font = Font(bold=True)
                impl_sheet[f'D{idx}'].font = Font(bold=True)
                impl_sheet[f'F{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 30 or idx == 30:
                impl_sheet[f'B{idx}'].number_format = '0.0'
                impl_sheet[f'C{idx}'].number_format = money_format
                impl_sheet[f'D{idx}'].number_format = money_format
                impl_sheet[f'F{idx}'].number_format = money_format

            if idx < 30:
                impl_sheet[f'E{idx}'].number_format = percent_format

        # Implementation Summary
        impl_sheet['A32'] = "Implementation Summary"
        impl_sheet['A32'].font = header_font

        summary_headers = ['Category', 'Our Cost', 'Client Price', 'Profit', 'Margin %']
        for idx, header in enumerate(summary_headers, 1):
            cell = impl_sheet.cell(row=33, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        summary_categories = [
            ('Project Setup', '=D9', '=F9'),
            ('System Configuration', '=D16', '=F16'),
            ('Custom Development', '=E23', '=G23'),
            ('Data Migration and Training', '=D30', '=F30'),
            ('Total Implementation', '=SUM(B34:B37)', '=SUM(C34:C37)')
        ]

        for idx, (category, cost, price) in enumerate(summary_categories, 34):
            impl_sheet[f'A{idx}'] = category
            impl_sheet[f'B{idx}'] = cost
            impl_sheet[f'C{idx}'] = price
            impl_sheet[f'D{idx}'] = f'=C{idx}-B{idx}'
            impl_sheet[f'E{idx}'] = f'=D{idx}/C{idx}'

            # Apply number formats
            impl_sheet[f'B{idx}'].number_format = money_format
            impl_sheet[f'C{idx}'].number_format = money_format
            impl_sheet[f'D{idx}'].number_format = money_format
            impl_sheet[f'E{idx}'].number_format = percent_format

            if idx == 38:  # Total row
                impl_sheet[f'A{idx}'].font = Font(bold=True)
                impl_sheet[f'B{idx}'].font = Font(bold=True)
                impl_sheet[f'C{idx}'].font = Font(bold=True)
                impl_sheet[f'D{idx}'].font = Font(bold=True)
                impl_sheet[f'E{idx}'].font = Font(bold=True)

        # Apply conditional formatting to margins
        for idx in range(34, 39):
            target_margin = 0.45
            if idx == 36:  # Custom Development - higher target
                target_margin = 0.50

            margin_rule = FormulaRule(
                formula=[f'E{idx}<{target_margin}'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
            )
            impl_sheet.conditional_formatting.add(f'E{idx}:E{idx}', margin_rule)

        # Adjust column widths
        for col in impl_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            impl_sheet.column_dimensions[column].width = adjusted_width

        return impl_sheet

        # ----------------------- Create Support & Maintenance Sheet -----------------------

    def create_support_sheet():
        support_sheet = wb.create_sheet("Support & Maintenance")
        add_navigation_bar(support_sheet)
        add_sheet_header(support_sheet, "Support and Maintenance")

        # Monthly Support Section
        support_sheet['A5'] = "Monthly Support"
        support_sheet['A5'].font = header_font

        # Set up column headers
        support_headers = ['Service Level', 'Base Monthly Fee', 'Additional Hours', 'Hourly Rate', 'Monthly Cost',
                           'Markup %', 'Monthly Price']
        for idx, header in enumerate(support_headers, 1):
            cell = support_sheet.cell(row=6, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Support services
        support_services = [
            ('Selected Support Tier', '=\'Client Info\'!B16',
             '=IF(A7="Basic",5,IF(A7="Standard",10,IF(A7="Premium",20,15)))',
             '=VLOOKUP("Support Specialist", Parameters!A7:C13, 3, FALSE)'),
            ('Hosting Management', 'N/A', 2, '=VLOOKUP("System Architect", Parameters!A7:C13, 3, FALSE)'),
            ('System Updates', 'N/A', 5, '=VLOOKUP("Developer", Parameters!A7:C13, 3, FALSE)'),
            ('Subtotal', '', '', '')
        ]

        for idx, (service, base_fee, add_hours, hourly_rate) in enumerate(support_services, 7):
            support_sheet[f'A{idx}'] = service

            if idx < 10:  # Regular rows
                support_sheet[f'B{idx}'] = base_fee
                if idx == 7:  # Support tier
                    support_sheet[
                        f'E{idx}'] = f'=IF(A7="Basic",\'License & Infrastructure\'!F30*0.12,IF(A7="Standard",\'License & Infrastructure\'!F30*0.18,IF(A7="Premium",\'License & Infrastructure\'!F30*0.25,2000)))*VLOOKUP(\'Client Info\'!B15,Parameters!A37:C40,3,FALSE)'
                else:
                    support_sheet[f'C{idx}'] = add_hours
                    support_sheet[f'D{idx}'] = hourly_rate
                    support_sheet[f'E{idx}'] = f'=C{idx}*D{idx}'

                support_sheet[f'F{idx}'] = 0.55  # 55% markup for support services
                support_sheet[f'G{idx}'] = f'=E{idx}*(1+F{idx})'
            else:  # Subtotal row
                support_sheet[f'E{idx}'] = '=SUM(E7:E9)'
                support_sheet[f'G{idx}'] = '=SUM(G7:G9)'
                support_sheet[f'A{idx}'].font = Font(bold=True)
                support_sheet[f'E{idx}'].font = Font(bold=True)
                support_sheet[f'G{idx}'].font = Font(bold=True)

            # Apply number formats
            if idx < 10 or idx == 10:
                if isinstance(hourly_rate, str) and hourly_rate.startswith('='):
                    support_sheet[f'D{idx}'].number_format = money_format
                support_sheet[f'E{idx}'].number_format = money_format
                support_sheet[f'G{idx}'].number_format = money_format

            if idx < 10:
                support_sheet[f'F{idx}'].number_format = percent_format

        # Support Summary
        support_sheet['A12'] = "Support Summary"
        support_sheet['A12'].font = header_font

        summary_headers = ['Category', 'Monthly Cost', 'Monthly Price', 'Annual Price', 'Profit', 'Margin %']
        for idx, header in enumerate(summary_headers, 1):
            cell = support_sheet.cell(row=13, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Support summary row
        support_sheet['A14'] = 'Total Support & Maintenance'
        support_sheet['B14'] = '=E10'
        support_sheet['C14'] = '=G10'
        support_sheet['D14'] = '=C14*12'
        support_sheet['E14'] = '=C14-B14'
        support_sheet['F14'] = '=E14/C14'

        # Apply formats
        support_sheet['B14'].number_format = money_format
        support_sheet['C14'].number_format = money_format
        support_sheet['D14'].number_format = money_format
        support_sheet['E14'].number_format = money_format
        support_sheet['F14'].number_format = percent_format

        support_sheet['A14'].font = Font(bold=True)
        support_sheet['B14'].font = Font(bold=True)
        support_sheet['C14'].font = Font(bold=True)
        support_sheet['D14'].font = Font(bold=True)
        support_sheet['E14'].font = Font(bold=True)
        support_sheet['F14'].font = Font(bold=True)

        # Apply conditional formatting
        support_margin_rule = FormulaRule(
            formula=['F14<Parameters!$B$32'],
            stopIfTrue=True,
            fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
        )
        support_sheet.conditional_formatting.add('F14:F14', support_margin_rule)

        # Support Overview - Key Metrics
        support_sheet['A16'] = "Support Metrics"
        support_sheet['A16'].font = header_font

        metrics = [
            ('Support as % of License Costs', '=C14/(\'License & Infrastructure\'!G31*12)'),
            ('Support Hours per User per Month', '=SUM(C7:C9)/\'Client Info\'!B25'),
            ('Average Support Cost per User', '=IF(\'Client Info\'!B25=0,0,C14/\'Client Info\'!B25)'),
            ('Support Tier Coverage',
             '=IF(A7="Basic","Normal business hours",IF(A7="Standard","Extended business hours",IF(A7="Premium","24/7 coverage","Custom coverage")))')
        ]

        for idx, (metric, formula) in enumerate(metrics, 17):
            support_sheet[f'A{idx}'] = metric
            support_sheet[f'B{idx}'] = formula

            if idx < 20:  # % and money formats
                if idx == 17:  # Percentage
                    support_sheet[f'B{idx}'].number_format = percent_format
                elif idx == 19:  # Money
                    support_sheet[f'B{idx}'].number_format = money_format
                else:
                    support_sheet[f'B{idx}'].number_format = '0.0'

        # Adjust column widths
        for col in support_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            support_sheet.column_dimensions[column].width = adjusted_width

        return support_sheet

        # ----------------------- Create Contract Summary Sheet -----------------------

    def create_contract_summary_sheet():
        summary_sheet = wb.create_sheet("Contract Summary")
        add_navigation_bar(summary_sheet)
        add_sheet_header(summary_sheet, "Contract Summary and Financial Analysis")

        # One-Time Costs Section
        summary_sheet['A5'] = "One-Time Costs"
        summary_sheet['A5'].font = header_font

        # Set up column headers
        onetime_headers = ['Category', 'Our Cost', 'Client Price', 'Profit', 'Margin %', 'Amortized Monthly']
        for idx, header in enumerate(onetime_headers, 1):
            cell = summary_sheet.cell(row=6, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # One-time cost categories
        onetime_categories = [
            ('Project Setup', '=Implementation!B34', '=Implementation!C34'),
            ('System Configuration', '=Implementation!B35', '=Implementation!C35'),
            ('Custom Development', '=Implementation!B36', '=Implementation!C36'),
            ('Data Migration and Training', '=Implementation!B37', '=Implementation!C37'),
            ('Setup Fees', 500, 1000),
            ('Total One-Time', '=SUM(B7:B11)', '=SUM(C7:C11)')
        ]

        for idx, (category, cost, price) in enumerate(onetime_categories, 7):
            summary_sheet[f'A{idx}'] = category
            summary_sheet[f'B{idx}'] = cost
            summary_sheet[f'C{idx}'] = price
            summary_sheet[f'D{idx}'] = f'=C{idx}-B{idx}'
            summary_sheet[f'E{idx}'] = f'=D{idx}/C{idx}'

            if idx < 12:  # Regular rows
                if idx == 11:  # Setup fees
                    summary_sheet[f'B{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                                fill_type="solid")
                    summary_sheet[f'C{idx}'].fill = PatternFill(start_color=COLOR_INPUT, end_color=COLOR_INPUT,
                                                                fill_type="solid")

                if idx == 12:  # Total row
                    summary_sheet[
                        f'F{idx}'] = f'=IF(OR(ISNA(VLOOKUP(\'Client Info\'!B13,Parameters!A44:D47,3,FALSE)),EXACT(VLOOKUP(\'Client Info\'!B13,Parameters!A44:D47,3,FALSE),"None")),C{idx},C{idx}/VALUE(LEFT(VLOOKUP(\'Client Info\'!B13,Parameters!A44:D47,3,FALSE),2)))'
                else:
                    summary_sheet[f'F{idx}'] = ''

            # Apply number formats
            summary_sheet[f'B{idx}'].number_format = money_format
            summary_sheet[f'C{idx}'].number_format = money_format
            summary_sheet[f'D{idx}'].number_format = money_format
            summary_sheet[f'E{idx}'].number_format = percent_format
            if idx == 12:
                summary_sheet[f'F{idx}'].number_format = money_format

            if idx == 12:  # Total row
                summary_sheet[f'A{idx}'].font = Font(bold=True)
                summary_sheet[f'B{idx}'].font = Font(bold=True)
                summary_sheet[f'C{idx}'].font = Font(bold=True)
                summary_sheet[f'D{idx}'].font = Font(bold=True)
                summary_sheet[f'E{idx}'].font = Font(bold=True)
                summary_sheet[f'F{idx}'].font = Font(bold=True)

        # Recurring Costs Section
        summary_sheet['A14'] = "Recurring Costs"
        summary_sheet['A14'].font = header_font

        # Set up column headers (same as one-time costs)
        for idx, header in enumerate(onetime_headers[:-1] + ['Annual Revenue'], 1):
            cell = summary_sheet.cell(row=15, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

            # Recurring cost categories
        recurring_categories = [
            ('Software Licenses', '=\'License & Infrastructure\'!D19', '=\'License & Infrastructure\'!F19',
             '=\'License & Infrastructure\'!G19'),
            ('Infrastructure', '=\'License & Infrastructure\'!D27', '=\'License & Infrastructure\'!F27',
             '=\'License & Infrastructure\'!G27'),
            ('Support & Maintenance', '=\'Support & Maintenance\'!B14', '=\'Support & Maintenance\'!C14',
             '=\'Support & Maintenance\'!C14'),
            ('One-Time Cost Amortization', '0', '=F12', '=F12'),
            ('Total Recurring', '=SUM(B16:B19)', '=SUM(C16:C19)', '=SUM(D16:D19)')
        ]

        for idx, (category, cost, price, monthly) in enumerate(recurring_categories, 16):
            summary_sheet[f'A{idx}'] = category
            summary_sheet[f'B{idx}'] = cost
            summary_sheet[f'C{idx}'] = price
            summary_sheet[f'D{idx}'] = f'=C{idx}-B{idx}'
            summary_sheet[f'E{idx}'] = f'=D{idx}/C{idx}'
            summary_sheet[f'F{idx}'] = monthly

            # Apply number formats
            summary_sheet[f'B{idx}'].number_format = money_format
            summary_sheet[f'C{idx}'].number_format = money_format
            summary_sheet[f'D{idx}'].number_format = money_format
            summary_sheet[f'E{idx}'].number_format = percent_format
            summary_sheet[f'F{idx}'].number_format = money_format

            if idx == 20:  # Total row
                summary_sheet[f'A{idx}'].font = Font(bold=True)
                summary_sheet[f'B{idx}'].font = Font(bold=True)
                summary_sheet[f'C{idx}'].font = Font(bold=True)
                summary_sheet[f'D{idx}'].font = Font(bold=True)
                summary_sheet[f'E{idx}'].font = Font(bold=True)
                summary_sheet[f'F{idx}'].font = Font(bold=True)

        # Calculate annual revenue
        summary_sheet['G20'] = '=F20*12'
        summary_sheet['G20'].number_format = money_format
        summary_sheet['G20'].font = Font(bold=True)

        # Contract Value Section
        summary_sheet['A22'] = "Total Contract Value"
        summary_sheet['A22'].font = header_font

        # Contract types
        contract_headers = ['Contract Term', 'Contract Years', 'Annual Revenue', 'Our Annual Cost',
                            'Total Contract Value', 'Total Profit', 'Overall Margin %']
        for idx, header in enumerate(contract_headers, 1):
            cell = summary_sheet.cell(row=23, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        contract_terms = [
            ('1st Year', 1, '=G20+C12', '=B20*12+B12'),
            ('2-Year Contract', 2,
             '=G20*(1-(IF(\'Client Info\'!B13="2-Year",VLOOKUP(\'Client Info\'!B13,Parameters!A44:D47,4,FALSE),0)))',
             '=B20*12'),
            ('3-Year Contract', 3,
             '=G20*(1-(IF(\'Client Info\'!B13="3-Year",VLOOKUP(\'Client Info\'!B13,Parameters!A44:D47,4,FALSE),0)))',
             '=B20*12'),
            ('Selected Contract', '=VALUE(LEFT(\'Client Info\'!B13,1))', '=IF(B27=1,C24,IF(B27=2,C25,C26))',
             '=IF(B27=1,D24,IF(B27=2,D25,D26))')
        ]

        for idx, (term, years, annual_rev, annual_cost) in enumerate(contract_terms, 24):
            summary_sheet[f'A{idx}'] = term
            summary_sheet[f'B{idx}'] = years
            summary_sheet[f'C{idx}'] = annual_rev
            summary_sheet[f'D{idx}'] = annual_cost
            summary_sheet[f'E{idx}'] = f'=IF(B{idx}=1,C{idx},C{idx}*B{idx})'
            summary_sheet[f'F{idx}'] = f'=E{idx}-IF(B{idx}=1,D{idx},D{idx}*B{idx})'
            summary_sheet[f'G{idx}'] = f'=F{idx}/E{idx}'

            # Apply number formats
            summary_sheet[f'C{idx}'].number_format = money_format
            summary_sheet[f'D{idx}'].number_format = money_format
            summary_sheet[f'E{idx}'].number_format = money_format
            summary_sheet[f'F{idx}'].number_format = money_format
            summary_sheet[f'G{idx}'].number_format = percent_format

            if idx == 27:  # Selected contract row
                summary_sheet[f'A{idx}'].font = Font(bold=True)
                summary_sheet[f'E{idx}'].font = Font(bold=True)
                summary_sheet[f'F{idx}'].font = Font(bold=True)
                summary_sheet[f'G{idx}'].font = Font(bold=True)

        # Apply conditional formatting
        summary_margin_rule = FormulaRule(
            formula=['G27<Parameters!$B$33'],
            stopIfTrue=True,
            fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
        )
        summary_sheet.conditional_formatting.add('G27:G27', summary_margin_rule)

        # Payment Schedule Section
        summary_sheet['A29'] = "Payment Schedule"
        summary_sheet['A29'].font = header_font

        payment_headers = ['Payment Type', 'Amount', 'Due Date', 'Notes']
        for idx, header in enumerate(payment_headers, 1):
            cell = summary_sheet.cell(row=30, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        payment_types = [
            ('Initial Payment', '=C12*0.5', 'Upon contract signing', '50% of implementation cost'),
            ('Second Payment', '=C12*0.5', 'Upon system deployment', '50% of implementation cost'),
            ('Monthly Recurring', '=F20', 'Monthly starting at go-live', 'Ongoing services and licenses')
        ]

        for idx, (payment, amount, due_date, notes) in enumerate(payment_types, 31):
            summary_sheet[f'A{idx}'] = payment
            summary_sheet[f'B{idx}'] = amount
            summary_sheet[f'C{idx}'] = due_date
            summary_sheet[f'D{idx}'] = notes

            # Apply number format
            summary_sheet[f'B{idx}'].number_format = money_format

        # Breakeven Analysis
        summary_sheet['A35'] = "Breakeven Analysis"
        summary_sheet['A35'].font = header_font

        breakeven_headers = ['Month', 'Cumulative Cost', 'Cumulative Revenue', 'Net Position']
        for idx, header in enumerate(breakeven_headers, 1):
            cell = summary_sheet.cell(row=36, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        for month in range(1, 13):  # 12 months
            row = 36 + month
            summary_sheet[f'A{row}'] = month

            if month == 1:
                summary_sheet[f'B{row}'] = '=B12*0.5+B20'  # Initial payment (50%) + first month recurring
                summary_sheet[f'C{row}'] = '=C12*0.5+F20'  # Initial payment (50%) + first month revenue
            elif month == 6:  # Assuming system deployment at month 6
                summary_sheet[
                    f'B{row}'] = f'=B{row - 1}+B12*0.5+B20'  # Previous month + second payment (50%) + monthly recurring
                summary_sheet[
                    f'C{row}'] = f'=C{row - 1}+C12*0.5+F20'  # Previous month + second payment (50%) + monthly revenue
            else:
                summary_sheet[f'B{row}'] = f'=B{row - 1}+B20'  # Previous month + monthly recurring
                summary_sheet[f'C{row}'] = f'=C{row - 1}+F20'  # Previous month + monthly revenue

            summary_sheet[f'D{row}'] = f'=C{row}-B{row}'  # Net position

            # Apply number formats
            summary_sheet[f'B{row}'].number_format = money_format
            summary_sheet[f'C{row}'].number_format = money_format
            summary_sheet[f'D{row}'].number_format = money_format

        # Add breakeven month calculation
        summary_sheet['A50'] = "Breakeven Month:"
        summary_sheet['A50'].font = Font(bold=True)
        summary_sheet['B50'] = '=MATCH(TRUE,D37:D48>=0,0)'
        summary_sheet['C50'] = 'months'

        # Apply conditional formatting to show breakeven
        for row in range(37, 49):
            breakeven_rule = FormulaRule(
                formula=[f'D{row}<0'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
            )
            summary_sheet.conditional_formatting.add(f'D{row}:D{row}', breakeven_rule)

            breakeven_rule_positive = FormulaRule(
                formula=[f'D{row}>=0'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_ABOVE_TARGET, end_color=COLOR_ABOVE_TARGET, fill_type="solid")
            )
            summary_sheet.conditional_formatting.add(f'D{row}:D{row}', breakeven_rule_positive)

        # Add breakeven chart
        breakeven_chart = LineChart()
        breakeven_chart.title = "Breakeven Analysis"
        breakeven_chart.style = 2
        breakeven_chart.y_axis.title = "Amount ($)"
        breakeven_chart.x_axis.title = "Month"
        breakeven_chart.height = 15
        breakeven_chart.width = 20

        # Add data
        costs = Reference(summary_sheet, min_col=2, min_row=36, max_row=48, max_col=2)
        revenue = Reference(summary_sheet, min_col=3, min_row=36, max_row=48, max_col=3)
        net = Reference(summary_sheet, min_col=4, min_row=36, max_row=48, max_col=4)

        breakeven_chart.add_data(costs, titles_from_data=False)
        breakeven_chart.add_data(revenue, titles_from_data=False)
        breakeven_chart.add_data(net, titles_from_data=False)

        # Add series names
        breakeven_chart.series[0].tx = SeriesLabel(v="Cumulative Cost")
        breakeven_chart.series[1].tx = SeriesLabel(v="Cumulative Revenue")
        breakeven_chart.series[2].tx = SeriesLabel(v="Net Position")

        # Set series markers and colors
        breakeven_chart.series[0].graphicalProperties.line.solidFill = "7030A0"  # Purple for costs
        breakeven_chart.series[1].graphicalProperties.line.solidFill = "00B050"  # Green for revenue
        breakeven_chart.series[2].graphicalProperties.line.solidFill = "0070C0"  # Blue for net

        for series in breakeven_chart.series:
            series.marker = Marker(symbol="circle", size=5)
            series.smooth = True

        # Add chart to worksheet
        summary_sheet.add_chart(breakeven_chart, "H35")

        # Adjust column widths
        for col in summary_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            summary_sheet.column_dimensions[column].width = adjusted_width

        return summary_sheet

        # ----------------------- Create Profitability Analysis Sheet -----------------------

    def create_profitability_sheet():
        profit_sheet = wb.create_sheet("Profitability Analysis")
        add_navigation_bar(profit_sheet)
        add_sheet_header(profit_sheet, "Profitability Analysis and Metrics")

        # Cost Breakdown Section
        profit_sheet['A5'] = "Cost Breakdown by Category"
        profit_sheet['A5'].font = header_font

        # Set up column headers
        profit_headers = ['Category', 'Target Margin', 'Revenue', 'Cost', 'Profit', 'Actual Margin', 'Variance']
        for idx, header in enumerate(profit_headers, 1):
            cell = profit_sheet.cell(row=6, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        # Cost categories
        categories = [
            ('Software Licenses', '=VLOOKUP("Software Licenses",Parameters!A28:B33,2,FALSE)',
             '=\'License & Infrastructure\'!F19', '=\'License & Infrastructure\'!D19'),
            ('Infrastructure', '=VLOOKUP("Infrastructure",Parameters!A28:B33,2,FALSE)',
             '=\'License & Infrastructure\'!F27', '=\'License & Infrastructure\'!D27'),
            ('Implementation', '=VLOOKUP("Implementation",Parameters!A28:B33,2,FALSE)',
             '=\'Implementation\'!C38', '=\'Implementation\'!B38'),
            ('Support & Maintenance', '=VLOOKUP("Support & Maintenance",Parameters!A28:B33,2,FALSE)',
             '=\'Support & Maintenance\'!D14', '=\'Support & Maintenance\'!B14*12'),
            ('Overall Solution', '=VLOOKUP("Overall Solution",Parameters!A28:B33,2,FALSE)',
             '=\'Contract Summary\'!E27', '=\'Contract Summary\'!E27-\'Contract Summary\'!F27')
        ]

        for idx, (category, target, revenue, cost) in enumerate(categories, 7):
            profit_sheet[f'A{idx}'] = category
            profit_sheet[f'B{idx}'] = target
            profit_sheet[f'C{idx}'] = revenue
            profit_sheet[f'D{idx}'] = cost
            profit_sheet[f'E{idx}'] = f'=C{idx}-D{idx}'
            profit_sheet[f'F{idx}'] = f'=E{idx}/C{idx}'
            profit_sheet[f'G{idx}'] = f'=F{idx}-B{idx}'

            # Apply number formats
            profit_sheet[f'B{idx}'].number_format = percent_format
            profit_sheet[f'C{idx}'].number_format = money_format
            profit_sheet[f'D{idx}'].number_format = money_format
            profit_sheet[f'E{idx}'].number_format = money_format
            profit_sheet[f'F{idx}'].number_format = percent_format
            profit_sheet[f'G{idx}'].number_format = percent_format

            if idx == 11:  # Overall row
                profit_sheet[f'A{idx}'].font = Font(bold=True)
                profit_sheet[f'E{idx}'].font = Font(bold=True)
                profit_sheet[f'F{idx}'].font = Font(bold=True)

        # Apply conditional formatting
        for idx in range(7, 12):
            margin_rule = FormulaRule(
                formula=[f'F{idx}<B{idx}'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
            )
            profit_sheet.conditional_formatting.add(f'F{idx}:F{idx}', margin_rule)

            variance_rule_negative = FormulaRule(
                formula=[f'G{idx}<0'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_BELOW_TARGET, end_color=COLOR_BELOW_TARGET, fill_type="solid")
            )
            profit_sheet.conditional_formatting.add(f'G{idx}:G{idx}', variance_rule_negative)

            variance_rule_positive = FormulaRule(
                formula=[f'G{idx}>=0'],
                stopIfTrue=True,
                fill=PatternFill(start_color=COLOR_ABOVE_TARGET, end_color=COLOR_ABOVE_TARGET, fill_type="solid")
            )
            profit_sheet.conditional_formatting.add(f'G{idx}:G{idx}', variance_rule_positive)

        # Add profitability chart
        pie_chart = PieChart()
        pie_chart.title = "Revenue Distribution"

        labels = Reference(profit_sheet, min_col=1, min_row=7, max_row=10)
        data = Reference(profit_sheet, min_col=3, min_row=7, max_row=10)
        pie_chart.add_data(data, titles_from_data=False)
        pie_chart.set_categories(labels)

        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True

        profit_sheet.add_chart(pie_chart, "A14")

        # Margin comparison chart
        margin_chart = BarChart()
        margin_chart.title = "Margin Comparison"
        margin_chart.style = 2
        margin_chart.type = "col"
        margin_chart.y_axis.title = "Margin %"
        margin_chart.x_axis.title = "Category"

        target_margins = Reference(profit_sheet, min_col=2, min_row=7, max_row=10)
        actual_margins = Reference(profit_sheet, min_col=6, min_row=7, max_row=10)
        categories = Reference(profit_sheet, min_col=1, min_row=7, max_row=10)

        margin_chart.add_data(target_margins, titles_from_data=False)
        margin_chart.add_data(actual_margins, titles_from_data=False)
        margin_chart.set_categories(categories)

        margin_chart.series[0].tx = SeriesLabel(v="Target Margin")
        margin_chart.series[1].tx = SeriesLabel(v="Actual Margin")

        profit_sheet.add_chart(margin_chart, "E14")

        # Key Financial Metrics
        profit_sheet['A30'] = "Key Financial Metrics"
        profit_sheet['A30'].font = header_font

        metrics = [
            ('Monthly Recurring Revenue (MRR)', '=\'Contract Summary\'!F20'),
            ('Annual Recurring Revenue (ARR)', '=\'Contract Summary\'!G20'),
            ('Total Contract Value (TCV)', '=\'Contract Summary\'!E27'),
            ('Average Revenue Per User (ARPU)',
             '=IF(\'Client Info\'!B25=0,0,\'Contract Summary\'!F20/\'Client Info\'!B25)'),
            ('Cost of Goods Sold (COGS)', '=D11'),
            ('Gross Profit', '=E11'),
            ('Gross Margin', '=F11'),
            ('Breakeven Point', '=\'Contract Summary\'!B50&" months"'),
            ('Customer Acquisition Cost (CAC)', '=\'Implementation\'!C38'),
            ('CAC Payback Period', '=\'Implementation\'!C38/E11*12&" months"')
        ]

        for idx, (metric, formula) in enumerate(metrics, 31):
            profit_sheet[f'A{idx}'] = metric
            profit_sheet[f'B{idx}'] = formula

            # Apply number formats
            if idx in [31, 32, 33, 34, 35, 36]:  # Money values
                profit_sheet[f'B{idx}'].number_format = money_format
            elif idx == 37:  # Percentage
                profit_sheet[f'B{idx}'].number_format = percent_format

        # Adjust column widths
        for col in profit_sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            profit_sheet.column_dimensions[column].width = adjusted_width

        return profit_sheet

        # ----------------------- Create Proposal Generator Sheet -----------------------

    def create_proposal_sheet():
        proposal_sheet = wb.create_sheet("Proposal Generator")
        add_navigation_bar(proposal_sheet)
        add_sheet_header(proposal_sheet, "Client Proposal Generator")

        # Client Information Section
        proposal_sheet['A5'] = "Client Proposal: WoodPro Geospatial Solution"
        proposal_sheet['A5'].font = Font(bold=True, size=14)

        proposal_sheet['A6'] = "Prepared for:"
        proposal_sheet['A6'].font = Font(bold=True)

        proposal_sheet['A7'] = '=\'Client Info\'!B6'  # Client name
        proposal_sheet['A8'] = "Date: " + pd.Timestamp.today().strftime('%B %d, %Y')

        # Solution Overview Section
        proposal_sheet['A10'] = "Solution Overview"
        proposal_sheet['A10'].font = header_font

        proposal_sheet[
            'A11'] = """This proposal outlines a comprehensive WoodPro implementation tailored for your organization, providing an enterprise-grade geospatial solution for efficient timber sales, inventory management, and field operations. The solution includes web-based and mobile access components to support both office and field staff."""
        proposal_sheet['A11'].alignment = Alignment(wrap_text=True)

        # Core Components Section
        proposal_sheet['A13'] = "Core Components:"
        proposal_sheet['A13'].font = Font(bold=True)

        core_components = [
            "• WoodPro Base Platform with ESRI Enterprise GIS integration",
            "• User Access Licenses for office and field personnel",
            "• Custom Integration with existing systems",
            "• Data Migration from legacy systems",
            "• User Training and implementation support",
            "• Ongoing Maintenance and Support"
        ]

        for idx, component in enumerate(core_components, 14):
            proposal_sheet[f'A{idx}'] = component

        # System Specifications
        proposal_sheet['A21'] = "System Specifications"
        proposal_sheet['A21'].font = header_font

        proposal_sheet['A22'] = "Infrastructure:"
        proposal_sheet['A22'].font = Font(bold=True)

        proposal_sheet['A23'] = "• ESRI GIS Enterprise v11.x with cores"
        proposal_sheet['A24'] = "• Cloud hosting on AWS infrastructure"
        proposal_sheet['A25'] = "• Secure data storage with automated backups"
        proposal_sheet['A26'] = "• Mobile-enabled field access"

        proposal_sheet['A28'] = "User Licenses:"
        proposal_sheet['A28'].font = Font(bold=True)

        license_text = [
            "• ='Client Info'!B21 Standard Partner (Creator) accounts for GIS administrators",
            "• ='Client Info'!B20 Basic Partner (Contributor) accounts for office staff",
            "• ='Client Info'!B23 WoodPro Editor licenses for management team",
            "• ='Client Info'!B22 Mobile Partner licenses for field staff",
            "• ='Client Info'!B24 WoodPro Viewer licenses for readonly access"
        ]

        for idx, text in enumerate(license_text, 29):
            proposal_sheet[f'A{idx}'] = text

        # Implementation Timeline
        proposal_sheet['A35'] = "Implementation Timeline"
        proposal_sheet['A35'].font = header_font

        timeline_headers = ['Phase', 'Activities', 'Duration']
        for idx, header in enumerate(timeline_headers, 1):
            cell = proposal_sheet.cell(row=36, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        timeline_phases = [
            ('Discovery & Planning', 'Requirements gathering, system design', '2 weeks'),
            ('Infrastructure Setup', 'Cloud environment configuration, ESRI setup', '2 weeks'),
            ('Core Configuration', 'WoodPro configuration, data layer setup',
             '=IF(\'Client Info\'!B15="Basic","2 weeks",IF(\'Client Info\'!B15="Standard","3 weeks",IF(\'Client Info\'!B15="Advanced","4 weeks","6 weeks")))'),
            ('Custom Development', 'Integration development, custom features',
             '=IF(\'Client Info\'!B15="Basic","2 weeks",IF(\'Client Info\'!B15="Standard","4 weeks",IF(\'Client Info\'!B15="Advanced","6 weeks","8 weeks")))'),
            ('Data Migration', 'Legacy data import and validation', '3 weeks'),
            ('Testing', 'System testing, user acceptance testing', '2 weeks'),
            ('Training', 'Admin and end-user training', '2 weeks'),
            ('Go-Live', 'Production deployment and support', '1 week'),
            ('Total Duration', '', '=\'Client Info\'!E6')
        ]

        for idx, (phase, activities, duration) in enumerate(timeline_phases, 37):
            proposal_sheet[f'A{idx}'] = phase
            proposal_sheet[f'B{idx}'] = activities
            proposal_sheet[f'C{idx}'] = duration

            if idx == 45:  # Total row
                proposal_sheet[f'A{idx}'].font = Font(bold=True)
                proposal_sheet[f'C{idx}'].font = Font(bold=True)

        # Investment Summary
        proposal_sheet['A47'] = "Investment Summary"
        proposal_sheet['A47'].font = header_font

        # One-time costs
        proposal_sheet['A48'] = "One-Time Costs"
        proposal_sheet['A48'].font = Font(bold=True)

        onetime_headers = ['Component', 'Details', 'Cost']
        for idx, header in enumerate(onetime_headers, 1):
            cell = proposal_sheet.cell(row=49, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        onetime_components = [
            ('Implementation Services', 'System setup, configuration, migration',
             '=\'Contract Summary\'!C7+\'Contract Summary\'!C8+\'Contract Summary\'!C10'),
            ('Custom Development', 'Accounting integration, custom reports', '=\'Contract Summary\'!C9'),
            ('Training', 'Admin and end-user training sessions', '=\'Implementation\'!F28'),
            ('Total One-Time', '', '=\'Contract Summary\'!C12')
        ]

        for idx, (component, details, cost) in enumerate(onetime_components, 50):
            proposal_sheet[f'A{idx}'] = component
            proposal_sheet[f'B{idx}'] = details
            proposal_sheet[f'C{idx}'] = cost

            # Apply money format
            proposal_sheet[f'C{idx}'].number_format = money_format

            if idx == 53:  # Total row
                proposal_sheet[f'A{idx}'].font = Font(bold=True)
                proposal_sheet[f'C{idx}'].font = Font(bold=True)

        # Annual recurring costs
        proposal_sheet['A55'] = "Annual Recurring Costs"
        proposal_sheet['A55'].font = Font(bold=True)

        recurring_headers = ['Component', 'Details', 'Annual Cost']
        for idx, header in enumerate(recurring_headers, 1):
            cell = proposal_sheet.cell(row=56, column=idx)
            cell.value = header
            cell.font = header_font
            cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
            cell.border = thin_border

        recurring_components = [
            ('ESRI Enterprise Licenses', 'Base platform + additional cores', '=\'License & Infrastructure\'!F9'),
            ('User Licenses', 'All license types (\'=Client Info\'!B25 total users)',
             '=\'License & Infrastructure\'!F19'),
            ('Infrastructure', 'AWS hosting, storage, security', '=\'License & Infrastructure\'!F27'),
            ('Support & Maintenance', '=\'Support & Maintenance\'!A7', '=\'Support & Maintenance\'!D14'),
            ('Total Annual', '', '=\'Contract Summary\'!G20')
        ]

        for idx, (component, details, cost) in enumerate(recurring_components, 57):
            proposal_sheet[f'A{idx}'] = component
            proposal_sheet[f'B{idx}'] = details
            proposal_sheet[f'C{idx}'] = cost

            # Apply money format
            proposal_sheet[f'C{idx}'].number_format = money_format

            if idx == 61:  # Total row
                proposal_sheet[f'A{idx}'].font = Font(bold=True)
                proposal_sheet[f'C{idx}'].font = Font(bold=True)
                # Payment Schedule
                proposal_sheet['A63'] = "Payment Schedule"
                proposal_sheet['A63'].font = Font(bold=True)

                schedule_headers = ['Payment', 'Amount', 'Due Date']
                for idx, header in enumerate(schedule_headers, 1):
                    cell = proposal_sheet.cell(row=64, column=idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                    cell.border = thin_border

                schedule_items = [
                    ('Initial Payment', '=\'Contract Summary\'!B31', '=\'Contract Summary\'!C31'),
                    ('Second Payment', '=\'Contract Summary\'!B32', '=\'Contract Summary\'!C32'),
                    ('Recurring Cost', '=\'Contract Summary\'!B33', '=\'Contract Summary\'!C33')
                ]

                for idx, (payment, amount, due_date) in enumerate(schedule_items, 65):
                    proposal_sheet[f'A{idx}'] = payment
                    proposal_sheet[f'B{idx}'] = amount
                    proposal_sheet[f'C{idx}'] = due_date

                    # Apply money format
                    proposal_sheet[f'B{idx}'].number_format = money_format

                # Return on Investment Analysis
                proposal_sheet['A69'] = "Return on Investment Analysis"
                proposal_sheet['A69'].font = header_font

                proposal_sheet[
                    'A70'] = "Based on industry benchmarks and client-specific operations, we project the following benefits:"

                benefit_headers = ['Benefit Area', 'Estimated Annual Savings']
                for idx, header in enumerate(benefit_headers, 1):
                    cell = proposal_sheet.cell(row=71, column=idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
                    cell.border = thin_border

                benefit_areas = [
                    ('Operational Efficiency', '$85,000 - $115,000'),
                    ('Reduced Travel Costs', '$25,000 - $35,000'),
                    ('Improved Inventory Management', '$45,000 - $65,000'),
                    ('Regulatory Compliance Management', '$30,000 - $40,000'),
                    ('Paper Reduction', '$10,000 - $15,000'),
                    ('Total Estimated Annual Benefit', '$195,000 - $270,000')
                ]

                for idx, (area, savings) in enumerate(benefit_areas, 72):
                    proposal_sheet[f'A{idx}'] = area
                    proposal_sheet[f'B{idx}'] = savings

                    if idx == 77:  # Total row
                        proposal_sheet[f'A{idx}'].font = Font(bold=True)
                        proposal_sheet[f'B{idx}'].font = Font(bold=True)

                proposal_sheet['A79'] = "Projected ROI: The solution is expected to pay for itself within "
                proposal_sheet[
                    'B79'] = "='Contract Summary'!B50 & \" - \" & ('Contract Summary'!B50 + 2) & \" months.\""

                proposal_sheet['A79'].font = Font(bold=True)
                proposal_sheet['B79'].font = Font(bold=True)

                # Support and Maintenance
                proposal_sheet['A81'] = "Support and Maintenance"
                proposal_sheet['A81'].font = header_font

                proposal_sheet['A82'] = '=\'Support & Maintenance\'!A7&" Support Package"'
                proposal_sheet['A82'].font = Font(bold=True)

                support_details = [
                    ('• Helpdesk support during business hours (8am-6pm ET)'),
                    ('• System monitoring and proactive maintenance'),
                    ('• Monthly system updates and patches'),
                    ('• Quarterly performance reviews'),
                    ('• Access to self-service knowledge base and training materials'),
                ]

                for idx, detail in enumerate(support_details, 83):
                    proposal_sheet[f'A{idx}'] = detail

                # Next Steps
                proposal_sheet['A90'] = "Next Steps"
                proposal_sheet['A90'].font = header_font

                next_steps = [
                    "1. Review and acceptance of proposal",
                    "2. Contract finalization",
                    "3. Project kickoff meeting",
                    "4. Detailed requirements gathering",
                    "5. Implementation commencement"
                ]

                for idx, step in enumerate(next_steps, 91):
                    proposal_sheet[f'A{idx}'] = step

                # Contact Information
                proposal_sheet['A98'] = "For questions regarding this proposal, please contact:"
                proposal_sheet['A99'] = "• Sales Contact: [Sales Manager Name]"
                proposal_sheet['A100'] = "• Technical Contact: [Technical Lead Name]"
                proposal_sheet['A101'] = "• Phone: (XXX) XXX-XXXX"
                proposal_sheet['A102'] = "• Email: sales@yourcompany.com"

                proposal_sheet['A104'] = "Valid for 60 days from " + pd.Timestamp.today().strftime('%B %d, %Y')
                proposal_sheet['A104'].font = Font(italic=True)

                # Set column widths
                proposal_sheet.column_dimensions['A'].width = 30
                proposal_sheet.column_dimensions['B'].width = 40
                proposal_sheet.column_dimensions['C'].width = 15

                return proposal_sheet

    # Create all sheets
    create_parameters_sheet()
    create_client_info_sheet()
    create_license_sheet()
    create_implementation_sheet()
    create_support_sheet()
    create_contract_summary_sheet()
    create_profitability_sheet()
    create_proposal_sheet()

    # Set the active sheet to Client Info
    wb.active = wb['Client Info']

    # Save the workbook

    filename = "Geospatial_Solution_Pricing_Calculator.xlsx"
    wb.save(filename)
    print(f"Excel workbook has been created successfully: {filename}")


    return filename
if __name__ == "__main__":
    create_geospatial_pricing_calculator()