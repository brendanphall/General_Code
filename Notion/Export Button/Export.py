"""
Notion Database to Excel Export Script

This script:
1. Connects to the Notion API
2. Retrieves data from a specified database
3. Processes the data into a pandas DataFrame
4. Formats and exports the data to an Excel file
5. Returns a download link or sends the file via email
"""

import os
import requests
import pandas as pd
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import json
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create Flask app for webhook endpoint
app = Flask(__name__)

# Notion API configuration - token is stored in environment variable for security
NOTION_TOKEN = os.environ.get("NOTION_TOKEN")
headers = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Content-Type": "application/json",
    "Notion-Version": "2022-06-28"  # API version to use
}

class NotionClient:
    """
    Client for interacting with the Notion API.
    Handles authentication and database queries.
    """
    def __init__(self, token):
        # Initialize with API token and set up headers for all requests
        self.token = token
        self.headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Notion-Version": "2022-06-28"
        }
    
    def query_database(self, database_id, filter_params=None):
        """
        Query a Notion database and return all results.
        Handles pagination automatically to retrieve all pages.
        """
        url = f"https://api.notion.com/v1/databases/{database_id}/query"
        
        # Initialize payload with any filter parameters
        payload = {}
        if filter_params:
            payload.update(filter_params)
        
        # Initialize variables for pagination
        all_results = []
        has_more = True
        next_cursor = None
        
        # Loop until all pages are retrieved
        while has_more:
            # Add cursor to payload if we're not on the first page
            if next_cursor:
                payload["start_cursor"] = next_cursor
                
            # Make API request
            response = requests.post(url, json=payload, headers=self.headers)
            data = response.json()
            
            # Check for errors
            if response.status_code != 200:
                raise Exception(f"Error querying database: {data}")
                
            # Add results to our collection
            all_results.extend(data.get("results", []))
            
            # Check if there are more pages
            has_more = data.get("has_more", False)
            next_cursor = data.get("next_cursor")
            
        return all_results
    
    def get_database_metadata(self, database_id):
        """
        Get database metadata including property schemas.
        This is needed to understand the structure of the database.
        """
        url = f"https://api.notion.com/v1/databases/{database_id}"
        response = requests.get(url, headers=self.headers)
        
        # Check for errors
        if response.status_code != 200:
            raise Exception(f"Error getting database metadata: {response.json()}")
            
        return response.json()

class PandasConverter:
    """
    Convert Notion data to pandas-friendly format.
    Handles the complex Notion property types and converts them to simple Python types.
    """
    
    @staticmethod
    def extract_property_value(prop_value, prop_type):
        """
        Extract the actual value from a Notion property based on its type.
        Notion properties are complex objects, this method extracts the actual value.
        """
        # Handle None values
        if prop_value is None:
            return None
            
        # Handle title properties (main title of a page)
        if prop_type == "title":
            if len(prop_value["title"]) > 0:
                return prop_value["title"][0]["plain_text"]
            return ""
            
        # Handle rich text properties (formatted text)
        elif prop_type == "rich_text":
            if len(prop_value["rich_text"]) > 0:
                return prop_value["rich_text"][0]["plain_text"]
            return ""
            
        # Handle number properties
        elif prop_type == "number":
            return prop_value["number"]
            
        # Handle select properties (single choice from options)
        elif prop_type == "select":
            if prop_value["select"]:
                return prop_value["select"]["name"]
            return None
            
        # Handle multi-select properties (multiple choices from options)
        elif prop_type == "multi_select":
            if prop_value["multi_select"]:
                return [item["name"] for item in prop_value["multi_select"]]
            return []
            
        # Handle date properties
        elif prop_type == "date":
            if prop_value["date"]:
                return prop_value["date"]["start"]
            return None
            
        # Handle checkbox properties (true/false)
        elif prop_type == "checkbox":
            return prop_value["checkbox"]
            
        # Handle URL properties
        elif prop_type == "url":
            return prop_value["url"]
            
        # Handle email properties
        elif prop_type == "email":
            return prop_value["email"]
            
        # Handle phone number properties
        elif prop_type == "phone_number":
            return prop_value["phone_number"]
            
        # Handle created time properties (when the page was created)
        elif prop_type == "created_time":
            return prop_value["created_time"]
            
        # Handle created by properties (who created the page)
        elif prop_type == "created_by":
            return prop_value["created_by"]["name"]
            
        # Handle last edited time properties (when the page was last edited)
        elif prop_type == "last_edited_time":
            return prop_value["last_edited_time"]
            
        # Handle last edited by properties (who last edited the page)
        elif prop_type == "last_edited_by":
            return prop_value["last_edited_by"]["name"]
            
        # Handle formula properties (computed values)
        elif prop_type == "formula":
            formula_type = prop_value["formula"]["type"]
            return prop_value["formula"][formula_type]
            
        # Handle relation properties (links to other pages)
        elif prop_type == "relation":
            return [item["id"] for item in prop_value["relation"]]
            
        # Handle people properties (users)
        elif prop_type == "people":
            return [person["name"] for person in prop_value["people"]]
            
        else:
            # Return a default for any other type
            return str(prop_value)

class ExcelFormatter:
    """
    Format and style Excel files.
    Applies professional styling to make the Excel output look good.
    """
    
    @staticmethod
    def format_excel(df, excel_path, sheet_name="Notion Data"):
        """
        Format the Excel file with styling.
        Adds headers, borders, colors, and adjusts column widths.
        """
        # Write basic DataFrame to Excel
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Access the workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define styles for headers and cells
            header_font = Font(bold=True, color="FFFFFF")  # White, bold text
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")  # Blue background
            border = Border(
                left=Side(style="thin"), 
                right=Side(style="thin"), 
                top=Side(style="thin"), 
                bottom=Side(style="thin")
            )
            
            # Format headers - apply styling to each header cell
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                
                # Adjust column width based on content length
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(
                    15, min(30, len(str(df.columns[col-1])) + 4)
                )
            
            # Format data cells - apply styling to each data cell
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    
                    # Alternate row coloring for better readability
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
            
            # Freeze top row so headers stay visible when scrolling
            worksheet.freeze_panes = "A2"
        
        return excel_path

class NotionToExcel:
    """
    Main class to handle Notion to Excel conversion.
    Orchestrates the entire export process.
    """
    
    def __init__(self, notion_token):
        # Initialize all required components
        self.notion_client = NotionClient(notion_token)
        self.pandas_converter = PandasConverter()
        self.excel_formatter = ExcelFormatter()
    
    def convert_to_dataframe(self, database_id):
        """
        Convert Notion database to pandas DataFrame.
        This is the core conversion function.
        """
        # Get database data and schema
        pages = self.notion_client.query_database(database_id)
        db_metadata = self.notion_client.get_database_metadata(database_id)
        properties_schema = db_metadata["properties"]
        
        # Extract property types for each field
        property_types = {
            prop_name: prop_schema["type"] 
            for prop_name, prop_schema in properties_schema.items()
        }
        
        # Build records for DataFrame - one record per page
        records = []
        for page in pages:
            record = {}
            for prop_name, prop_schema in properties_schema.items():
                prop_type = property_types[prop_name]
                if prop_name in page["properties"]:
                    # Extract the value using the appropriate converter
                    value = self.pandas_converter.extract_property_value(
                        page["properties"][prop_name], prop_type
                    )
                    record[prop_name] = value
                else:
                    # Handle missing properties
                    record[prop_name] = None
            records.append(record)
        
        # Create DataFrame from the records
        df = pd.DataFrame(records)
        return df
    
    def export_to_excel(self, database_id, file_path=None):
        """
        Export Notion database to formatted Excel file.
        This is the main entry point for the export process.
        """
        # Convert Notion data to DataFrame
        df = self.convert_to_dataframe(database_id)
        
        # Generate file path if not provided
        if file_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_path = f"notion_export_{timestamp}.xlsx"
        
        # Format and save Excel file
        self.excel_formatter.format_excel(df, file_path)
        return file_path

# Flask webhook endpoint - this is the API endpoint that clients will call
@app.route('/export-notion-to-excel', methods=['POST'])
def export_notion_to_excel():
    """
    API endpoint to handle export requests.
    Accepts a database ID and returns an Excel file.
    """
    try:
        # Get database ID from request
        data = request.json
        database_id = data.get('database_id')
        
        # Validate input
        if not database_id:
            return jsonify({"error": "Database ID is required"}), 400
        
        # Create temp file for the export
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_file_path = temp_file.name
        temp_file.close()
        
        # Export to Excel
        exporter = NotionToExcel(NOTION_TOKEN)
        excel_file = exporter.export_to_excel(database_id, temp_file_path)
        
        # You could implement different return methods here:
        # 1. Return file for download
        # 2. Save to cloud storage and return link
        # 3. Email the file as attachment
        
        # For now, we'll just return the file directly
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=f"notion_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # Handle any errors that occur during the export process
        return jsonify({"error": str(e)}), 500

# Main entry point - run the Flask app
if __name__ == '__main__':
    # Get port from environment variable or use default
    port = int(os.environ.get('PORT', 5000))
    # Run the app on all network interfaces
    app.run(host='0.0.0.0', port=port)